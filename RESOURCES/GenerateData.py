from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
fake_data = Faker()


# Generating test data
def generate_data():
    ws.cell(row=1, column=1).value = "Email"
    ws.cell(row=1, column=2).value = "Password"
    ws.cell(row=1, column=3).value = "Used"
    for i in range(2, 101):
        ws.cell(row=i, column=1).value = f"{fake_data.first_name()}.{fake_data.last_name()}@gmail.com"
        ws.cell(row=i, column=2).value = fake_data.bothify(text="??#?#?#?#?#")
    # wb.save("RESOURCES/TestData.xlsx")
    wb.save(filename="RESOURCES/TestData.xlsx")





