from openpyxl import load_workbook
import GenerateData

wb = load_workbook(filename="RESOURCES/TestData.xlsx")
ws = wb.active
number_of_rows = 9
counter_rows = 0


# Choosing not used row in the Test Data.xlsx
def not_used():
    for i in range(2, number_of_rows - 1):
        while ws.cell(row=i, column=3).value == None:
            row_id = i
            return row_id
            break


row_id = not_used()


# Assigning email address to the variable
def new_user_data_username():
    username = ws.cell(row=row_id, column=1).value
    return username


# Assigning password to the variable
def new_user_data_password():
    user_password = ws.cell(row=row_id, column=2).value
    return user_password


# Marking the row which was used in testing with "Yes"
def add_used():
    ws.cell(row=row_id, column=3).value = "Yes"
    wb.save("RESOURCES/TestData.xlsx")


# Assigning text of the modal's(PROBLEMS LOGGING IN?) body to the variable
def problems_logging_in_body_text():
    file_to_read = "RESOURCES/EN/LOGIN_MODAL_Problems_Logging_In_BODY_EN.txt"
    file = open(file_to_read, "r")
    file_data = file.read()
    return file_data
    file.close()


# Assigning text of the modal's(PROBLEMS LOGGING IN?) header to the variable
def problems_logging_in_header_text():
    file_to_read = "RESOURCES/EN/LOGIN_MODAL_Problems_Logging_In_HEADER_EN.txt"
    file = open(file_to_read, "r")
    file_data = file.read()
    return file_data
    file.close()


# Assigning text of the error to the variable
def incorrect_username_or_password_error_text():
    file_to_read = "RESOURCES/EN/LOGIN_Incorrect_username_or_password_ERROR_EN.txt"
    file = open(file_to_read, "r")
    file_data = file.read()
    return file_data
    file.close()


# Assigning text of the modal's(DON'T HAVE AN ACCOUNT?) body to the variable
def dont_have_an_account_body_text():
    file_to_read = "RESOURCES/EN/LOGIN_MODAL_Don't_have_an_account_BODY_EN.txt"
    file = open(file_to_read, "r")
    file_data = file.read()
    return file_data
    file.close()


# Assigning text of the modal's(DON'T HAVE AN ACCOUNT?) header to the variable
def dont_have_an_account_header_text():
    file_to_read = "RESOURCES/EN/LOGIN_MODAL_Don't_have_an_account_HEADER_EN.txt"
    file = open(file_to_read, "r")
    file_data = file.read()
    return file_data
    file.close()


# Assigning text of the modal's(Privacy Statement) body to the variable
def privacy_statement_body_text():
    file_to_read = "RESOURCES/EN/LOGIN_MODAL_Privacy_Statement_BODY_EN.txt"
    file = open(file_to_read, "r")
    file_data = file.read()
    return file_data
    file.close()


# Assigning text of the error to the variable - Suomi language
def incorrect_username_or_password_error_text_suomi():
    file_to_read = "RESOURCES/FI/LOGIN_Incorrect_username_or_password_ERROR_FI.txt"
    file = open(file_to_read, "r", encoding="utf-8")
    file_data = file.read()
    return file_data
    file.close()
